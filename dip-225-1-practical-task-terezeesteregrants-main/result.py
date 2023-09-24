from openpyxl import Workbook, load_workbook
wb = load_workbook('tests/test1.xlsx')
ws = wb.active
total = 0
max_row = ws.max_row
for row in range(2,max_row+1):
    rate = (ws['B'+ str(row)].value)
    hours = (ws['C'+ str(row)].value)
    
    if (type(hours) != str and type(rate) != str):
        salary = float(rate) * hours
        if (salary > 3000):
            total += 1
        ws['D' + str(row)] = salary

 
print(total)

wb.save('tests/test1.xlsx')
wb.close()